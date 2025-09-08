# This script takes a "limitlesstcg.com" deck link and converts it into a sheets/Excel file.
# Makes it convenient for creating a checklist for required cards
import openpyxl
import requests
import pandas as pd
from lxml import html
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def get_deck(url):
    """
    Grabs deck list from the given limitless TCG url
    :param url: URL of desired deck
    :return: dict containing card list
    """
    page = requests.get(url)
    tree = html.fromstring(page.content)

    columns = tree.xpath('//div[contains(@class, "decklist-column")]')
    deck_data = []
    for column in columns:
        if len(column) == 0:
            continue
        else:
            card_counts = column.xpath('.//span[contains(@class, "card-count")]/text()')
            card_names = column.xpath('.//span[contains(@class, "card-name")]/text()')
            expansions = column.xpath('.//div[contains(@class, "decklist-card")]/@data-set')
            prices = column.xpath('.//a[contains(@class, "card-price usd")]/text()')

            column_data = convert_deck_to_dict(card_counts, card_names, expansions, prices)
            deck_data.append(column_data)

    return {
        "url": url,
        "cards": deck_data
    }

def convert_deck_to_dict(counts, names, expansions, prices):
    """
    Converts list of cards to dict
    :param counts: List of numbers for each type of card
    :param names: List of names for each type of card
    :param expansions: List of expansions for each type of card
    :return: dict containing deck list
    """
    new_dict = dict()
    for count, name, expansion, price in zip(counts, names, expansions, prices):
        try:
            new_dict[f"{name.strip()} ({expansion.strip()})"] = {
                "Count": int(count.strip()),
                "Price": price.strip()
            }
        except ValueError:
            continue
    return new_dict

def export_dict_into_xlsx(data, deck_name="Deck"):
    """
    Converts dict to xlsx file
    :param data: Dict containing deck list
    :param deck_name: Name of deck
    :return: xlsx file
    """
    rows = []
    # for key, values in data.items():
    for category in data["cards"]:
        for key, value in category.items():
            rows.append([key, value["Price"], value["Count"], 0, f"=IF(OR(C{len(rows) + 2}=\"\", D{len(rows) + 2}=\"\"), \"\", C{len(rows) + 2}-D{len(rows) + 2})"])
        rows.append([None, None, None, None])

    df = pd.DataFrame(rows, columns=["Card Name", "Market Price (USD)", "Quantity Needed", "Amount Have", "Remaining"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = deck_name

    # Append DataFrame rows to Excel sheet
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    last_row = len(df) + 1  # +1 for header row, +1 to account for 1-based indexing
    ws.append(["# remaining cards", "", "", f"=SUM(E2:E{last_row - 1})"])
    ws.append([f"{data['url']}", "", "", ""])

    # Adjust column width based on data size
    max_width=20

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = min(max_width, max(max_length, len(str(cell.value))))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Define conditional formatting
    green_fill = PatternFill(start_color="b7e1cd", end_color="b7e1cd", fill_type="solid")
    red_fill = PatternFill(start_color="ea9999", end_color="ea9999", fill_type="solid")

    # Apply conditional formatting only if the cell is not empty
    ws.conditional_formatting.add(
        f"E2:E{last_row - 1}",
        FormulaRule(formula=["AND(ISNUMBER(E2), E2<=0)"], fill=green_fill)
    )
    ws.conditional_formatting.add(
        f"E2:E{last_row - 1}",
        FormulaRule(formula=["AND(ISNUMBER(E2), E2>0)"], fill=red_fill)
    )

    # Highlight the sum cell in yellow
    yellow_fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
    ws[f"D{last_row+1}"].fill = yellow_fill

    # Error check for whether file can be closed
    try:
        wb.save(f'{deck_name}.xlsx')
    except PermissionError as e:
        raise e
    else:
        print(f"{deck_name}.xlsx created.")

def convert(url: str, filename: str):
    # deck_name = sys.argv[2].strip() if len(sys.argv) > 2 else "Deck"
    deck = get_deck(url)
    export_dict_into_xlsx(deck, filename)

if __name__ == '__main__':
    url = "https://limitlesstcg.com/decks/list/18922"
    convert(url, "DeckList.xlsx")

